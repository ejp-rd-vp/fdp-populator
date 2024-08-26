import Config
import openpyxl
from resource_classes import VPOrganisation, VPBiobank, VPPatientregistry, VPDataset, VPDistribution, VPDataService

class VPTemplateReader:
    """
    NOTE: this class is based on the following specification as of November 10 2023:
    <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>
    """

    separator = "|"
    row = []
    keys = []

    def getval(self, key):
        """
        This method returns a value
        from a row based on the key

        :return value
        """
        return self.row[self.keys[key]].value

    def getvals(self, key):
        """
        This method returns multiple values
        from a row based on the key and separator

        :return values
        
        """
        entry = self.row[self.keys[key]].value
        if type(entry) == str:
            return [value.strip() for value in entry.split(self.separator)]
        return []

    def check_template_version(self):
        """
        This method checks whether the Excel template is the expected version

        :return: nothing
        """
        print("Checking sheet names...")
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        expected_sheets = ['Organisation', 'ContactPoint', 'Biobank', 
                           'PatientRegistry', 'Guideline', 'Dataset', 
                           'Distribution', 'DataService', 'Catalog']

        sheet_exists = [sheet in wb.sheetnames for sheet in expected_sheets]
        if False in sheet_exists:
            raise SystemError("A sheet in the Excel template is missing. The sheet could be a different version.")
        
        print("Excel template contains expected sheets.")

    def get_organisations(self):
        """
        This method creates organisation objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of organisations
        """
        # Prepare reading
        print("Reading organisation sheet...")
        expected_column_names = ['Title', 'Description', 'LandingPage',
                                  'Logo', 'Location', 'Identifier']
        keys = dict(zip(expected_column_names, range(0, len(expected_column_names))))

        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['Organisation']

        # Loop over rows of excel sheet
        first_row = True
        organisations = {}
        for row in ws:
            # Check header
            if first_row:
                first_row=False
                column_names = [cell.value for cell in row]
                if column_names != expected_column_names:
                    raise SystemError("Column names do not match in the organisation sheet")
                continue

            # Read row if it exists
            if row[keys["Title"]].value != None:
                # Create organisation object and add to organisation dictionary
                self.row = row
                self.keys = keys
                organisation = VPOrganisation.VPOrganisation(
                    parent_url=Config.CATALOG_URL,
                    title=self.getval("Title"),
                    description=self.getval("Description"),
                    location=self.getval("Location"),
                    pages=self.getvals("LandingPage"),
                    logo=self.getval("Logo"),
                    identifier=self.getval("Identifier"))
                organisations[organisation.TITLE] = organisation
                if Config.DEBUG: print(vars(organisation))

        return organisations

    def get_biobanks(self):
        """
        This method creates biobank objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of biobanks
        """
        # Prepare reading
        print("Reading biobank sheet...")
        expected_column_names = ['License', 'Title', 'Description', 'Theme',
                        'Publisher', 'ContactPoint', 'PersonalData',
                        'PopulationCoverage', 'Language', 'AccessRights',
                        'LandingPage', 'Distribution', 'VPConnection',
                        'ODRL Policy', 'Keyword', 'Logo', 'Identifier',
                        'Issued', 'Modified', 'Version', 'ConformsTo', None]
        keys = dict(zip(expected_column_names, range(0, len(expected_column_names))))
        
        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['Biobank']
        
        # Loop over rows of excel sheet
        first_row = True
        biobanks = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                column_names = [cell.value for cell in row]
                if column_names != expected_column_names:
                    raise SystemError("Column names do not match in the biobank sheet")
                continue

            # Read row if it exists
            if row[keys["Title"]].value != None:
                # Create biobank object and add to biobank dictionary if it is a biobank
                self.row = row
                self.keys = keys
                biobank = VPBiobank.VPBiobank(
                    parent_url=Config.CATALOG_URL,
                    license=self.getval("License"),
                    title=self.getval("Title"),
                    description=self.getval("Description"),
                    theme=self.getvals("Theme"),
                    publisher=self.getval("Publisher"),
                    contactpoint=self.getval("ContactPoint"),
                    language=self.getval("Language"),
                    personaldata=self.getval("PersonalData"),
                    conformsto=self.getval("ConformsTo"),
                    vpconnection=self.getvals("VPConnection"),
                    keyword=self.getvals("Keyword"),
                    logo=self.getval("Logo"),
                    haspolicy=self.getvals("ODRL Policy"),
                    identifier=self.getval("Identifier"),
                    issued=self.getval("Issued"),
                    modified=self.getval("Modified"),
                    version=self.getval("Version"),
                    accessrights=self.getval("AccessRights"),
                    landingpage=self.getval("LandingPage"),
                    distribution=self.getval("Distribution"),
                    populationcoverage=self.getval("PopulationCoverage"))

                biobanks[biobank.TITLE] = biobank
                if Config.DEBUG: print(vars(biobank))

        return biobanks

    def get_patientregistries(self):
        """
        This method creates patient registry objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of patientregistries
        """
        # Prepare reading
        print("Reading patient registry sheet...")
        expected_column_names = ['License', 'Title', 'Description',
                'Theme', 'Publisher', 'ContactPoint', 'PersonalData',
                'PopulationCoverage', 'Language', 'AccessRights',
            	'LandingPage', 'Distribution', 'VPConnection',
                'ODRL Policy', 'Keyword', 'Logo', 'Identifier',
                'Issued', 'Modified', 'Version', 'ConformsTo', None]
        keys = dict(zip(expected_column_names, range(0, len(expected_column_names))))

        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['PatientRegistry']
        
        # Loop over rows of excel sheet
        first_row = True
        patientregistries = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                column_names = [cell.value for cell in row]
                if column_names != expected_column_names:
                    raise SystemError("Column names do not match in the patient registry sheet")
                continue

            # Read row if it exists
            if row[0].value != None:
                # Create patient registry object and add to patientregistry dictionary if it is a patientregistry
                self.row = row
                self.keys = keys
                patientregistry = VPPatientregistry.VPPatientRegistry(
                    parent_url=Config.CATALOG_URL,
                    license=self.getval("License"),
                    title=self.getval("Title"),
                    description=self.getval("Description"),
                    theme=self.getvals("Theme"),
                    publisher=self.getval("Publisher"),
                    contactpoint=self.getval("ContactPoint"),
                    language=self.getval("Language"),
                    personaldata=self.getval("PersonalData"),
                    conformsto=self.getval("ConformsTo"),
                    vpconnection=self.getvals("VPConnection"),
                    keyword=self.getvals("Keyword"),
                    logo=self.getval("Logo"),
                    haspolicy=self.getvals("ODRL Policy"),
                    identifier=self.getval("Identifier"),
                    issued=self.getval("Issued"),
                    modified=self.getval("Modified"),
                    version=self.getval("Version"),
                    accessrights=self.getval("AccessRights"),
                    landingpage=self.getval("LandingPage"),
                    distribution=self.getval("Distribution"),
                    populationcoverage=self.getval("PopulationCoverage"))
                patientregistries[patientregistry.TITLE] = patientregistry
                if Config.DEBUG: print(vars(patientregistry))

        return patientregistries

    def get_datasets(self):
        """
        This method creates dataset objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of datasets
        """
        print("Reading dataset sheet...")
        expected_column_names = ['License', 'Title', 'Description', 'Theme',
                        'Publisher', 'ContactPoint', 'PersonalData',
                        'PopulationCoverage', 'Language', 'AccessRights',
                        'LandingPage', 'Distribution', 'VPConnection',
                        'ODRL Policy', 'Keyword', 'Logo', 'Identifier',
                        'Issued', 'Modified', 'Version', 'ConformsTo', None]
        keys = dict(zip(expected_column_names, range(0, len(expected_column_names))))

        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['Dataset']
        
        # Loop over rows of excel sheet
        first_row = True
        datasets = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                column_names = [cell.value for cell in row]
                if column_names != expected_column_names:
                    raise SystemError("Column names do not match in the dataset sheet")
                continue

            # Read row if it exists
            if row[0].value != None:
                # Create dataset object and add to dataset dictionary
                self.row = row
                self.keys = keys
                dataset = VPDataset.VPDataset(
                    parent_url=Config.CATALOG_URL,
                    license=self.getval("License"),
                    title=self.getval("Title"),
                    description=self.getval("Description"),
                    theme=self.getvals("Theme"),
                    publisher=self.getval("Publisher"),
                    contactpoint=self.getval("ContactPoint"),
                    language=self.getval("Language"),
                    personaldata=self.getval("PersonalData"),
                    conformsto=self.getval("ConformsTo"),
                    vpconnection=self.getvals("VPConnection"),
                    keyword=self.getvals("Keyword"),
                    logo=self.getval("Logo"),
                    haspolicy=self.getvals("ODRL Policy"),
                    identifier=self.getval("Identifier"),
                    issued=self.getval("Issued"),
                    modified=self.getval("Modified"),
                    version=self.getval("Version"),
                    accessrights=self.getvals("AccessRights"),
                    landingpage=self.getvals("LandingPage"),
                    distribution=self.getval("Distribution"))
                datasets[dataset.TITLE] = dataset
                if Config.DEBUG: print(vars(dataset))

        return datasets

    def get_distributions(self):
        """
        This method creates distribution objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of distributions
        """
        print("Reading distribution sheet...")
        expected_column_names = ['License', 'Title', 'Description',
            'Publisher', 'Version', 'AccessRights', 'ODRLPolicy',
            'MediaType', 'IsPartOf', 'Type', 'AccessService', 'Dataset Title']
        keys = dict(zip(expected_column_names, range(0, len(expected_column_names))))

        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['Distribution']
        
        # Loop over rows of excel sheet
        first_row = True
        distributions = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                column_names = [cell.value for cell in row]
                if column_names != expected_column_names:
                    raise SystemError("Column names do not match in the distribution sheet")
                continue

            # Read row if it exists
            if row[0].value != None:
                # Create distribution object and add to distribution dictionary
                self.row = row
                self.keys = keys
                distribution = VPDistribution.VPDistribution(
                    parent_url=None,
                    license=self.getval("License"),
                    title=self.getval("Title"),
                    description=self.getval("Description"),
                    publisher=self.getval("Publisher"),
                    version=self.getval("Version"),
                    accessrights=self.getval("AccessRights"),
                    haspolicy=self.getvals("ODRLPolicy"),
                    mediatype=self.getval("MediaType"),
                    ispartof=self.getvals("IsPartOf"),
                    accessurl=None,
                    downloadurl=None,
                    accessservice=self.getval("AccessService"),
                    conformsto=None,
                    dataset_title=self.getval("Dataset Title")
                )
                distributions[distribution.TITLE] = distribution
                if Config.DEBUG: print(vars(distribution))

        return distributions
    
    def get_dataservices(self):
        """
        This method creates dataservice objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of dataservices
        """
        print("Reading dataservice sheet...")
        expected_column_names = ['License', 'Type', 'Title',
            'Description', 'PersonalData', 'Publisher', 'Theme',
            'Language', 'ContactPoint', 'PopulationCoverage',
            'AccessRights', 'ConformsTo', 'EndpointDescription',
            'EndpointURL', 'LandingPage', 'VPConnection',
            'ODRLPolicy', 'Logo', 'ServesDataset', 'Keyword',
            'Identifier', 'Issued', 'Modified', 'Version', 
            'ConformsTo', None]
        keys = dict(zip(expected_column_names, range(0, len(expected_column_names))))

        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['DataService']
        
        # Loop over rows of excel sheet
        first_row = True
        dataservices = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                column_names = [cell.value for cell in row]
                if column_names != expected_column_names:
                    raise SystemError("Column names do not match in the dataservice sheet")
                continue

            # Read row if it exists
            if row[0].value != None:
                # Create dataservice object and add to dataservice dictionary
                self.row = row
                self.keys = keys
                dataservice = VPDataService.VPDataService(
                    parent_url=Config.CATALOG_URL,
                    license=self.getval("License"),
                    title=self.getval("Title"),
                    description=self.getval("Description"),
                    theme=self.getvals("Theme"),
                    publisher=self.getval("Publisher"),
                    contactpoint=self.getval("ContactPoint"),
                    language=self.getval("Language"),
                    personaldata=self.getval("PersonalData"),
                    conformsto=self.getval("ConformsTo"),
                    vpconnection=self.getvals("VPConnection"),
                    keyword=self.getvals("Keyword"),
                    logo=self.getval("Logo"),
                    haspolicy=self.getvals("ODRLPolicy"),
                    identifier=self.getval("Identifier"),
                    issued=self.getval("Issued"),
                    modified=self.getval("Modified"),
                    version=self.getval("Version"),
                    accessrights=self.getval("AccessRights"),
                    landingpage=self.getval("LandingPage"),
                    otype=self.getval("Type"),
                    servesdataset=self.getvals("ServesDataset"),
                    endpointurl=self.getval("EndpointURL"),
                    endpointdescription=self.getvals("EndpointDescription")
                    )
                dataservices[dataservice.TITLE] = dataservice
                if Config.DEBUG: print(vars(dataservice))

        return dataservices